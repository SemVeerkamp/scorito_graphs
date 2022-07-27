import numpy as np
import pandas as pd
import sys
import openpyxl
import matplotlib.pyplot as plt
import plotly.express as px


link = "C:/Users\semve\Desktop\Overig\scorito site\scorito\data.xlsx"

wb = openpyxl.load_workbook(link)

size = len(wb.sheetnames)
sheetnames = wb.sheetnames
del(sheetnames[0])

names = ["Sem", "Max", "Paul", "Jan", "Peter", "Wendela", "Gerard", "Ralph", "Maarten", "Loys", "Per", "Loeka", "Henriette"]


results = {}
for i in range(1, size):
    df = pd.read_excel(link, sheet_name=[i], index_col=0)
    results[sheetnames[i-1]] = df[i]

Standen = {}
for i in sheetnames:
    a = results[i]["Stand"]
    b = a.tolist()
    Standen[i] = b



Dagzeges = {}
for i in sheetnames:
    a = results[i]["Dagzeges"]
    b = a.tolist()
    Dagzeges[i] = b

Scorito = {}
for i in sheetnames:
    a = results[i]["Scorito-score"]
    b = a.tolist()
    Scorito[i] = b

Punten = {}
for i in sheetnames:
    a = results[i]["Punten"]
    b = a.tolist()
    Punten[i] = b


Overzicht={}
Winners=[]
Winner_scores=[]
Winner_scorito=[]
for sheet in sheetnames:
    edition = []
    for place in range(len(Standen[sheet])):
        name = Standen[sheet][place]
        punten = Punten[sheet][place]
        scorito = Scorito[sheet][place]
        dagzege = Dagzeges[sheet][place]
        list = [place+1, name, punten, scorito, dagzege]
        edition.append(list)
    Overzicht[sheet]=edition
    Winners.append(Standen[sheet][0])
    Winner_scores.append(Punten[sheet][0])
    Winner_scorito.append(Scorito[sheet][0])

Posities = {}
Scores = {}
Scorito_scores = {}
for name in names:
    positie=[]
    score=[]
    scorito_score=[]
    for sheet in sheetnames:
        plek = np.nan
        a = np.nan
        b = np.nan
        for place in range(len(Standen[sheet])):
            if name == Standen[sheet][place]:
                plek=place+1
                positie.append(plek)
                a=Punten[sheet][place]
                score.append(a)
                b=Scorito[sheet][place]
                scorito_score.append(b)
        if name not in Standen[sheet]:
            positie.append(plek)
            score.append(a)
            scorito_score.append(b)
    Posities[name]=positie
    Scores[name]=score
    Scorito_scores[name]=scorito_score

names = ["Sem", "Max", "Paul", "Jan", "Peter", "Wendela", "Gerard", "Ralph", "Maarten", "Loys", "Per", "Loeka", "Henriette"]

data = {"Wedstrijd": sheetnames,
        "Sem": Scores["Sem"],
        "Max": Scores["Max"],
        "Paul": Scores["Paul"],
        "Jan": Scores["Jan"],
        "Peter": Scores["Peter"],
        "Wendela": Scores["Wendela"],
        "Gerard": Scores["Gerard"],
        "Ralph": Scores["Ralph"],
        "Maarten": Scores["Maarten"],
        "Loys": Scores["Loys"],
        "Per": Scores["Per"],
        "Loeka": Scores["Loeka"],
        "Henriette": Scores["Henriette"],
        }
df_punten=pd.DataFrame(data)

data_scorito = {"Wedstrijd": sheetnames,
        "Sem": Scorito_scores["Sem"],
        "Max": Scorito_scores["Max"],
        "Paul": Scorito_scores["Paul"],
        "Jan": Scorito_scores["Jan"],
        "Peter": Scorito_scores["Peter"],
        "Wendela": Scorito_scores["Wendela"],
        "Gerard": Scorito_scores["Gerard"],
        "Ralph": Scorito_scores["Ralph"],
        "Maarten": Scorito_scores["Maarten"],
        "Loys": Scorito_scores["Loys"],
        "Per": Scorito_scores["Per"],
        "Loeka": Scorito_scores["Loeka"],
        "Henriette": Scorito_scores["Henriette"],
        }
df_scorito=pd.DataFrame(data_scorito)



fig = px.line(df_punten, x='Wedstrijd', y=df_punten.columns[1:],markers=True)

fig = px.line(df_scorito, x='Wedstrijd', y=df_scorito.columns[1:],markers=True)

fig.write_html('first_figure.html', auto_open=True)